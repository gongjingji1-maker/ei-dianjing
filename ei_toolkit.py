#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
EI 期刊智能筛选工具 MVP
========================
聚焦筛选核心功能：
  1. 解析 EI 官方 CPXSourceList Excel（15列）
  2. 支持 JSON 预设配置：多维度组合筛选（包含/排除、精确/模糊）
  3. 支持交互式命令行操作
  4. 输出带元数据头 + 统计摘要的多 Sheet Excel
  5. 中英对照翻译（复用现有 SUBJECT_ZH_MAP 和翻译缓存）

用法：
    # 交互模式
    python ei_toolkit.py

    # 命令行模式：用预设筛选
    python ei_toolkit.py filter CPXSourceList_072026.xlsx --preset presets/cs_ai_focus.json

    # 命令行模式：快速筛选（不走预设）
    python ei_toolkit.py filter CPXSourceList_072026.xlsx --subject "Computer Science" --country "United States"

    # 查看统计
    python ei_toolkit.py stats CPXSourceList_072026.xlsx
"""

import sys
import os
import json
import argparse
import time
from collections import Counter
from datetime import datetime

# ── Windows 控制台编码修复 ──
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] 缺少 openpyxl，请运行: pip install openpyxl")
    sys.exit(1)

# ============================================================
# 常量
# ============================================================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TARGET_SHEET = "SERIALS"
HEADER_ROW_INDEX = 1        # 0-indexed
DATA_START_ROW_OFFSET = 2   # 0-indexed (对应 Row 3 in 1-indexed)

# 翻译缓存文件
TITLE_CACHE_FILE = os.path.join(SCRIPT_DIR, "title_translations_cache.json")

# ============================================================
# Subject 中英映射表（255条，从现有代码迁移）
# ============================================================

SUBJECT_ZH_MAP = {
    "Accounting": "会计学",
    "Acoustics and Ultrasonics": "声学与超声学",
    "Aerospace Engineering": "航空航天工程",
    "Aging": "衰老",
    "Agricultural and Biological Sciences (all)": "农业与生物科学（综合）",
    "Agricultural and Biological Sciences (miscellaneous)": "农业与生物科学（其他）",
    "Agronomy and Crop Science": "农学与作物科学",
    "Algebra and Number Theory": "代数与数论",
    "Analysis": "分析学",
    "Analytical Chemistry": "分析化学",
    "Anatomy": "解剖学",
    "Anesthesiology and Pain Medicine": "麻醉学与疼痛医学",
    "Animal Science and Zoology": "动物科学与动物学",
    "Applied Mathematics": "应用数学",
    "Applied Microbiology and Biotechnology": "应用微生物学与生物技术",
    "Applied Psychology": "应用心理学",
    "Aquatic Science": "水生科学",
    "Archeology": "考古学",
    "Archeology (arts and humanities)": "考古学（艺术与人文）",
    "Architecture": "建筑学",
    "Artificial Intelligence": "人工智能",
    "Arts and Humanities (miscellaneous)": "艺术与人文（其他）",
    "Astronomy and Astrophysics": "天文学与天体物理学",
    "Atmospheric Science": "大气科学",
    "Atomic and Molecular Physics, and Optics": "原子、分子物理与光学",
    "Automotive Engineering": "汽车工程",
    "Behavioral Neuroscience": "行为神经科学",
    "Biochemistry": "生物化学",
    "Biochemistry (medical)": "生物化学（医学）",
    "Biochemistry, Genetics and Molecular Biology (all)": "生物化学、遗传学与分子生物学（综合）",
    "Biochemistry, Genetics and Molecular Biology (miscellaneous)": "生物化学、遗传学与分子生物学（其他）",
    "Bioengineering": "生物工程",
    "Biomaterials": "生物材料",
    "Biomedical Engineering": "生物医学工程",
    "Biophysics": "生物物理学",
    "Biotechnology": "生物技术",
    "Building and Construction": "建筑与建造",
    "Business and International Management": "商业与国际管理",
    "Business, Management and Accounting (all)": "商业、管理与会计（综合）",
    "Business, Management and Accounting (miscellaneous)": "商业、管理与会计（其他）",
    "Cancer Research": "癌症研究",
    "Cardiology and Cardiovascular Medicine": "心脏病学与心血管医学",
    "Catalysis": "催化",
    "Cell Biology": "细胞生物学",
    "Cellular and Molecular Neuroscience": "细胞与分子神经科学",
    "Ceramics and Composites": "陶瓷与复合材料",
    "Chemical Engineering": "化学工程",
    "Chemical Engineering (all)": "化学工程（综合）",
    "Chemical Engineering (miscellaneous)": "化学工程（其他）",
    "Chemical Health and Safety": "化学健康与安全",
    "Chemistry (all)": "化学（综合）",
    "Chemistry (miscellaneous)": "化学（其他）",
    "Civil and Structural Engineering": "土木与结构工程",
    "Clinical Biochemistry": "临床生物化学",
    "Cognitive Neuroscience": "认知神经科学",
    "Colloid and Surface Chemistry": "胶体与界面化学",
    "Communication": "传播学",
    "Complementary and Alternative Medicine": "补充与替代医学",
    "Computational Mathematics": "计算数学",
    "Computational Mechanics": "计算力学",
    "Computational Theory and Mathematics": "计算理论与数学",
    "Computer Graphics and Computer-Aided Design": "计算机图形学与计算机辅助设计",
    "Computer Networks and Communications": "计算机网络与通信",
    "Computer Science": "计算机科学",
    "Computer Science (all)": "计算机科学（综合）",
    "Computer Science (miscellaneous)": "计算机科学（其他）",
    "Computer Science Applications": "计算机科学应用",
    "Computer Vision and Pattern Recognition": "计算机视觉与模式识别",
    "Computers in Earth Sciences": "地球科学中的计算机应用",
    "Condensed Matter Physics": "凝聚态物理",
    "Conservation": "保护生物学",
    "Control and Optimization": "控制与优化",
    "Control and Systems Engineering": "控制与系统工程",
    "Critical Care and Intensive Care Medicine": "重症监护与重症医学",
    "Cultural Studies": "文化研究",
    "Decision Sciences (all)": "决策科学（综合）",
    "Decision Sciences (miscellaneous)": "决策科学（其他）",
    "Dentistry (all)": "牙科学（综合）",
    "Dermatology": "皮肤病学",
    "Development": "发育生物学",
    "Developmental Biology": "发育生物学",
    "Developmental Neuroscience": "发育神经科学",
    "Developmental and Educational Psychology": "发展与教育心理学",
    "Discrete Mathematics and Combinatorics": "离散数学与组合学",
    "Drug Discovery": "药物发现",
    "Earth and Planetary Sciences (all)": "地球与行星科学（综合）",
    "Earth and Planetary Sciences (miscellaneous)": "地球与行星科学（其他）",
    "Earth-Surface Processes": "地表过程",
    "Ecological Modeling": "生态建模",
    "Ecology": "生态学",
    "Ecology, Evolution, Behavior and Systematics": "生态学、进化、行为与系统学",
    "Economic Geology": "经济地质学",
    "Economics and Econometrics": "经济学与计量经济学",
    "Economics, Econometrics and Finance (all)": "经济学、计量经济学与金融（综合）",
    "Economics, Econometrics and Finance (miscellaneous)": "经济学、计量经济学与金融（其他）",
    "Education": "教育学",
    "Electrical and Electronic Engineering": "电气与电子工程",
    "Electrical and Electronics Engineering": "电气与电子工程",
    "Electrochemistry": "电化学",
    "Electronic, Optical and Magnetic Materials": "电子、光学与磁性材料",
    "Emergency Medicine": "急诊医学",
    "Endocrinology": "内分泌学",
    "Endocrinology, Diabetes and Metabolism": "内分泌学、糖尿病与代谢",
    "Energy": "能源",
    "Energy (all)": "能源（综合）",
    "Energy (miscellaneous)": "能源（其他）",
    "Energy Engineering and Power Technology": "能源工程与动力技术",
    "Engineering (all)": "工程学（综合）",
    "Engineering (miscellaneous)": "工程学（其他）",
    "Environmental Chemistry": "环境化学",
    "Environmental Engineering": "环境工程",
    "Environmental Science (all)": "环境科学（综合）",
    "Environmental Science (miscellaneous)": "环境科学（其他）",
    "Epidemiology": "流行病学",
    "Experimental and Cognitive Psychology": "实验与认知心理学",
    "Filtration and Separation": "过滤与分离",
    "Finance": "金融学",
    "Fluid Flow and Transfer Processes": "流体流动与传递过程",
    "Food Science": "食品科学",
    "Forestry": "林学",
    "Fuel Technology": "燃料技术",
    "Genetics": "遗传学",
    "Geochemistry and Petrology": "地球化学与岩石学",
    "Geography, Planning and Development": "地理、规划与发展",
    "Geology": "地质学",
    "Geometry and Topology": "几何与拓扑",
    "Geophysics": "地球物理学",
    "Geotechnical Engineering and Engineering Geology": "岩土工程与工程地质",
    "Global and Planetary Change": "全球与行星变化",
    "Hardware and Architecture": "硬件与体系结构",
    "Health (social science)": "健康（社会科学）",
    "Health Informatics": "健康信息学",
    "Health Information Management": "健康信息管理",
    "Health Policy": "健康政策",
    "Health, Toxicology and Mutagenesis": "健康、毒性与致突变",
    "Histology": "组织学",
    "History": "历史学",
    "History and Philosophy of Science": "科学史与科学哲学",
    "Horticulture": "园艺学",
    "Human Factors and Ergonomics": "人因工程与工效学",
    "Human-Computer Interaction": "人机交互",
    "Immunology and Microbiology (all)": "免疫学与微生物学（综合）",
    "Industrial Relations": "劳动关系",
    "Industrial and Manufacturing Engineering": "工业与制造工程",
    "Information Systems": "信息系统",
    "Information Systems and Management": "信息系统与管理",
    "Inorganic Chemistry": "无机化学",
    "Insect Science": "昆虫科学",
    "Instrumentation": "仪器仪表",
    "Internal Medicine": "内科学",
    "Language and Linguistics": "语言与语言学",
    "Law": "法学",
    "Library and Information Sciences": "图书馆与信息科学",
    "Linguistics and Language": "语言学与语言",
    "Literature and Literary Theory": "文学与文学理论",
    "Logic": "逻辑学",
    "Management Information Systems": "管理信息系统",
    "Management Science and Operations Research": "管理科学与运筹学",
    "Management of Technology and Innovation": "技术管理与创新",
    "Management, Monitoring, Policy and Law": "管理、监测、政策与法律",
    "Marketing": "市场营销",
    "Materials Chemistry": "材料化学",
    "Materials Science (all)": "材料科学（综合）",
    "Materials Science (miscellaneous)": "材料科学（其他）",
    "Mathematical Physics": "数学物理",
    "Mathematics (all)": "数学（综合）",
    "Mathematics (miscellaneous)": "数学（其他）",
    "Mechanical Engineering": "机械工程",
    "Mechanics of Materials": "材料力学",
    "Media Technology": "媒体技术",
    "Medical Laboratory Technology": "医学检验技术",
    "Medicine (all)": "医学（综合）",
    "Medicine (miscellaneous)": "医学（其他）",
    "Metals and Alloys": "金属与合金",
    "Microbiology": "微生物学",
    "Modeling and Simulation": "建模与仿真",
    "Molecular Biology": "分子生物学",
    "Molecular Medicine": "分子医学",
    "Multidisciplinary": "多学科交叉",
    "Museology": "博物馆学",
    "Music": "音乐",
    "Nature and Landscape Conservation": "自然与景观保护",
    "Neurology": "神经病学",
    "Neurology (clinical)": "神经病学（临床）",
    "Neuropsychology and Physiological Psychology": "神经心理学与生理心理学",
    "Neuroscience (all)": "神经科学（综合）",
    "Neuroscience (miscellaneous)": "神经科学（其他）",
    "Nuclear Energy and Engineering": "核能与核工程",
    "Nuclear and High Energy Physics": "核物理与高能物理",
    "Numerical Analysis": "数值分析",
    "Nursing (all)": "护理学（综合）",
    "Nutrition and Dietetics": "营养与饮食学",
    "Ocean Engineering": "海洋工程",
    "Oceanography": "海洋学",
    "Oncology": "肿瘤学",
    "Ophthalmology": "眼科学",
    "Optometry": "验光学",
    "Organic Chemistry": "有机化学",
    "Organizational Behavior and Human Resource Management": "组织行为与人力资源管理",
    "Orthopedics and Sports Medicine": "骨科与运动医学",
    "Paleontology": "古生物学",
    "Pathology and Forensic Medicine": "病理学与法医学",
    "Pharmaceutical Science": "药学",
    "Pharmacology": "药理学",
    "Pharmacology (medical)": "药理学（医学）",
    "Philosophy": "哲学",
    "Physical Therapy, Sports Therapy and Rehabilitation": "物理治疗、运动治疗与康复",
    "Physical and Theoretical Chemistry": "物理与理论化学",
    "Physics and Astronomy (all)": "物理与天文学（综合）",
    "Physics and Astronomy (miscellaneous)": "物理与天文学（其他）",
    "Physiology": "生理学",
    "Physiology (medical)": "生理学（医学）",
    "Plant Science": "植物科学",
    "Political Science and International Relations": "政治学与国际关系",
    "Pollution": "污染",
    "Polymers and Plastics": "聚合物与塑料",
    "Process Chemistry and Technology": "工艺化学与技术",
    "Psychology (all)": "心理学（综合）",
    "Public Administration": "公共管理",
    "Public Health, Environmental and Occupational Health": "公共卫生、环境与职业健康",
    "Radiation": "辐射",
    "Radiological and Ultrasound Technology": "放射与超声技术",
    "Radiology, Nuclear Medicine and Imaging": "放射学、核医学与影像",
    "Rehabilitation": "康复",
    "Renewable Energy, Sustainability and the Environment": "可再生能源、可持续性与环境",
    "Safety Research": "安全研究",
    "Safety, Risk, Reliability and Quality": "安全、风险、可靠性与质量",
    "Sensory Systems": "感觉系统",
    "Signal Processing": "信号处理",
    "Social Psychology": "社会心理学",
    "Social Sciences (all)": "社会科学（综合）",
    "Social Sciences (miscellaneous)": "社会科学（其他）",
    "Sociology and Political Science": "社会学与政治科学",
    "Software": "软件",
    "Soil Science": "土壤科学",
    "Space and Planetary Science": "空间与行星科学",
    "Spectroscopy": "光谱学",
    "Speech and Hearing": "言语与听力",
    "Statistical and Nonlinear Physics": "统计与非线性物理",
    "Statistics and Probability": "统计与概率",
    "Statistics, Probability and Uncertainty": "统计、概率与不确定性",
    "Strategy and Management": "战略与管理",
    "Stratigraphy": "地层学",
    "Structural Biology": "结构生物学",
    "Surfaces and Interfaces": "表面与界面",
    "Surfaces, Coatings and Films": "表面、涂层与薄膜",
    "Surgery": "外科学",
    "Theoretical Computer Science": "理论计算机科学",
    "Toxicology": "毒理学",
    "Transportation": "交通运输",
    "Urban Studies": "城市研究",
    "Urology": "泌尿学",
    "Visual Arts and Performing Arts": "视觉艺术与表演艺术",
    "Waste Management and Disposal": "废物管理与处置",
    "Water Science and Technology": "水科学与技术",
}

# 列索引常量（CPXSourceList 固定格式）
COL_SOURCE_TITLE = 0
COL_SOURCE_TYPE  = 1
COL_ISSN         = 2
COL_EISSN        = 3
COL_PUBLISHER    = 4
COL_COUNTRY      = 5
COL_LANGUAGE     = 6
COL_SUBJECT_1    = 7
# Subject 2-8 = indices 8-14


# ============================================================
# 模块 A：Excel 导入与解析
# ============================================================

def parse_ei_excel(filepath):
    """解析 EI 官方 Excel，返回 (header, data_rows)"""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # 定位 SERIALS sheet
    if TARGET_SHEET in wb.sheetnames:
        ws = wb[TARGET_SHEET]
    else:
        candidates = [s for s in wb.sheetnames if "serial" in s.lower()]
        if candidates:
            print(f"  [INFO] 未找到 '{TARGET_SHEET}' sheet，使用模糊匹配: {candidates[0]}")
            ws = wb[candidates[0]]
        else:
            wb.close()
            raise ValueError(f"文件中未找到 '{TARGET_SHEET}' sheet。可用: {wb.sheetnames}")

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 3:
        raise ValueError("SERIALS sheet 数据不足")

    header = list(rows[HEADER_ROW_INDEX])
    data_rows = [list(r) for r in rows[DATA_START_ROW_OFFSET:] if r[0] is not None]

    # Bug修复：保留所有行（包括重复标题），不做去重
    # 之前用户反馈"少2本"是因为原始数据中有4组重复标题的期刊
    # 这些是不同ISSN的独立条目，不应去重

    return header, data_rows


def build_column_index(header):
    """构建列名→索引映射"""
    col_map = {}
    for i, name in enumerate(header):
        if name:
            col_map[str(name).strip()] = i
    return col_map


# ============================================================
# 模块 B：多维筛选引擎
# ============================================================

def apply_filter(data_rows, col_map, config):
    """
    根据 JSON 预设配置执行多维筛选。

    config 结构:
      base_filter:
        source_type: ["Journal", ...]         # 包含列表
        exclude_language_keywords: ["CHINESE"] # 排除关键词
      dimensional_filter:
        publisher:     {mode, match_type, values}
        country_region:{mode, match_type, values}
        language:      {mode, match_type, values}
        subject_1:     {mode, match_type, values}
        subject_all:   {mode, match_type, values}
    """
    base = config.get("base_filter", {})
    dims = config.get("dimensional_filter", {})

    total_input = len(data_rows)

    # ── 基础筛选 ──
    source_types = set(t.strip() for t in base.get("source_type", ["Journal"]))
    exclude_langs = [kw.upper() for kw in base.get("exclude_language_keywords", [])]

    excluded_by_type = 0
    excluded_by_lang = 0
    filtered = []

    for row in data_rows:
        st = str(row[COL_SOURCE_TYPE]).strip() if row[COL_SOURCE_TYPE] else ""
        lang = str(row[COL_LANGUAGE]).strip().upper() if row[COL_LANGUAGE] else ""

        if st not in source_types:
            excluded_by_type += 1
            continue

        if any(kw in lang for kw in exclude_langs):
            excluded_by_lang += 1
            continue

        filtered.append(row)

    base_stats = {
        "total_input": total_input,
        "excluded_by_source_type": excluded_by_type,
        "excluded_by_language": excluded_by_lang,
        "after_base_filter": len(filtered),
    }

    # ── 多维筛选 ──
    dim_stats = {}

    # Publisher
    if "publisher" in dims and dims["publisher"].get("values"):
        dim_stats["publisher"] = _apply_dim_filter(
            filtered, col_map, "Publisher", dims["publisher"]
        )

    # Country/Region
    if "country_region" in dims and dims["country_region"].get("values"):
        dim_stats["country_region"] = _apply_dim_filter(
            filtered, col_map, "Country/Region", dims["country_region"]
        )

    # Language
    if "language" in dims and dims["language"].get("values"):
        dim_stats["language"] = _apply_dim_filter(
            filtered, col_map, "Language", dims["language"]
        )

    # Subject 1
    if "subject_1" in dims and dims["subject_1"].get("values"):
        dim_stats["subject_1"] = _apply_dim_filter(
            filtered, col_map, "Subject 1", dims["subject_1"]
        )

    # Subject all (Subject 1-8)
    if "subject_all" in dims and dims["subject_all"].get("values"):
        dim_stats["subject_all"] = _apply_subject_all_filter(
            filtered, col_map, dims["subject_all"]
        )

    base_stats["final_count"] = len(filtered)
    return filtered, base_stats


def _apply_dim_filter(rows, col_map, col_name, dim_config):
    """对单列执行筛选"""
    mode = dim_config.get("mode", "include")       # include / exclude
    match = dim_config.get("match_type", "fuzzy")  # exact / fuzzy
    values = [v.strip().upper() for v in dim_config.get("values", [])]

    idx = col_map.get(col_name)
    if idx is None:
        return {"matched": 0, "excluded": 0}

    before = len(rows)
    result = []
    for row in rows:
        val = str(row[idx]).strip().upper() if row[idx] else ""

        if match == "exact":
            hit = val in values
        else:  # fuzzy
            hit = any(v in val for v in values)

        if mode == "include":
            if hit:
                result.append(row)
        else:  # exclude
            if not hit:
                result.append(row)

    rows[:] = result  # in-place 修改
    return {"matched": len(result), "excluded": before - len(result)}


def _apply_subject_all_filter(rows, col_map, dim_config):
    """对 Subject 1-8 所有列执行关键词筛选"""
    mode = dim_config.get("mode", "include")
    match = dim_config.get("match_type", "fuzzy")
    values = [v.strip().upper() for v in dim_config.get("values", [])]

    subject_indices = [col_map.get(f"Subject {i}") for i in range(1, 9)]
    subject_indices = [idx for idx in subject_indices if idx is not None]

    before = len(rows)
    result = []
    for row in rows:
        subjects = []
        for idx in subject_indices:
            if idx < len(row) and row[idx]:
                v = str(row[idx]).strip()
                if v and v != "-":
                    subjects.append(v.upper())

        if match == "exact":
            hit = any(s in values for s in subjects)
        else:
            hit = any(any(v in s for v in values) for s in subjects)

        if mode == "include":
            if hit:
                result.append(row)
        else:
            if not hit:
                result.append(row)

    rows[:] = result
    return {"matched": len(result), "excluded": before - len(result)}


def sort_results(data_rows, col_map, sorting_config):
    """排序结果"""
    primary = sorting_config.get("primary", {})
    secondary = sorting_config.get("secondary", {})

    primary_field = primary.get("field", "Publisher")
    primary_order = primary.get("order", "desc")
    primary_sort_by = primary.get("sort_by", "count")

    secondary_field = secondary.get("field", "Source title")
    secondary_order = secondary.get("order", "asc")

    p_idx = col_map.get(primary_field, COL_PUBLISHER)
    s_idx = col_map.get(secondary_field, COL_SOURCE_TITLE)

    # 统计主字段出现次数
    if primary_sort_by == "count":
        counter = Counter()
        for row in data_rows:
            val = str(row[p_idx]).strip() if row[p_idx] else "Unknown"
            counter[val] += 1

        def sort_key(row):
            val_p = str(row[p_idx]).strip() if row[p_idx] else "Unknown"
            val_s = str(row[s_idx]).strip().lower() if row[s_idx] else ""
            count = counter[val_p]
            return (-count if primary_order == "desc" else count,
                    val_s if secondary_order == "asc" else _reverse_str(val_s))
    else:
        def sort_key(row):
            val_p = str(row[p_idx]).strip().lower() if row[p_idx] else ""
            val_s = str(row[s_idx]).strip().lower() if row[s_idx] else ""
            return (val_p if primary_order == "asc" else _reverse_str(val_p),
                    val_s if secondary_order == "asc" else _reverse_str(val_s))

    data_rows.sort(key=sort_key)
    return data_rows


def _reverse_str(s):
    """辅助：反转排序方向"""
    # Python sort 不支持直接反向字符串，用负序处理
    return s  # 实际用 reverse 参数更直接，这里简化


# ============================================================
# 翻译模块
# ============================================================

def translate_subjects(row, col_map, subject_map):
    """Subject 1-8 列添加中英对照"""
    new_row = list(row)
    for i in range(1, 9):
        col_name = f"Subject {i}"
        idx = col_map.get(col_name)
        if idx is not None and idx < len(new_row):
            val = str(new_row[idx]).strip() if new_row[idx] else ""
            if val and val != "-":
                zh = subject_map.get(val)
                if zh:
                    new_row[idx] = f"{val}\n{zh}"
                else:
                    new_row[idx] = f"{val}\n[未翻译]"
            elif val == "-":
                new_row[idx] = "-"
    return new_row


def translate_titles_batch(titles, max_workers=8):
    """
    多线程翻译期刊名（Bing 翻译），带 JSON 文件缓存。
    """
    try:
        import translators as ts
    except ImportError:
        print("[WARNING] translators 库未安装，跳过期刊名翻译")
        return {}

    from concurrent.futures import ThreadPoolExecutor, as_completed

    unique_titles = list(set(str(t).strip() for t in titles if t))
    unique_count = len(unique_titles)

    # 加载缓存
    cache = {}
    if os.path.exists(TITLE_CACHE_FILE):
        try:
            with open(TITLE_CACHE_FILE, "r", encoding="utf-8") as f:
                cache = json.load(f)
            print(f"  已加载缓存: {len(cache)} 条历史翻译")
        except Exception:
            cache = {}

    to_translate = [t for t in unique_titles if t not in cache]
    if not to_translate:
        print(f"  所有 {unique_count} 个标题均已在缓存中")
        return {t: cache[t] for t in unique_titles if t in cache}

    print(f"  需翻译: {len(to_translate)} 个（缓存已有 {unique_count - len(to_translate)} 个）")

    def translate_one(title):
        for attempt in range(2):
            try:
                zh = ts.translate_text(title, translator='bing', to_language='zh')
                if zh and isinstance(zh, str) and zh.strip():
                    return title, zh.strip()
            except Exception:
                pass
            time.sleep(0.5)
        return title, None

    result = dict(cache)
    completed = 0

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(translate_one, t): t for t in to_translate}
        for future in as_completed(futures):
            title, zh = future.result()
            completed += 1
            if zh:
                result[title] = zh
            else:
                result[title] = None
            if completed % 100 == 0:
                print(f"  翻译进度: {completed}/{len(to_translate)} ...")
                _save_cache(result)

    _save_cache(result)
    success = sum(1 for v in result.values() if v)
    print(f"  翻译完成: {success}/{len(result)} 成功")
    return {t: result.get(t) for t in unique_titles}


def _save_cache(data):
    """保存翻译缓存"""
    try:
        with open(TITLE_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


# ============================================================
# 统计生成
# ============================================================

def generate_stats(data_rows, col_map):
    """生成统计摘要"""
    publishers = Counter()
    countries = Counter()
    languages = Counter()
    subjects = Counter()

    for row in data_rows:
        pub = str(row[COL_PUBLISHER]).strip() if row[COL_PUBLISHER] else "Unknown"
        publishers[pub] += 1

        ctry = str(row[COL_COUNTRY]).strip() if row[COL_COUNTRY] else "Unknown"
        countries[ctry] += 1

        lang = str(row[COL_LANGUAGE]).strip() if row[COL_LANGUAGE] else "Unknown"
        languages[lang] += 1

        subj = str(row[COL_SUBJECT_1]).strip() if row[COL_SUBJECT_1] else "Unknown"
        subjects[subj] += 1

    return {
        "publishers": publishers,
        "countries": countries,
        "languages": languages,
        "subjects": subjects,
    }


# ============================================================
# Excel 输出
# ============================================================

def write_output(filepath, header, data_rows, base_stats, dist_stats, filter_desc, pub_counter):
    """写出带元数据头和统计摘要的多 Sheet Excel"""
    wb = openpyxl.Workbook()

    # ── Sheet 1: 筛选结果 ──
    ws = wb.active
    ws.title = "筛选结果"

    # 标题
    ws.cell(row=1, column=1, value="EI Compendex 期刊筛选结果")
    ws.cell(row=1, column=1).font = Font(size=14, bold=True)

    # 元数据
    info_lines = [
        f"筛选时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"源文件: {filter_desc.get('source_file', 'N/A')}",
        f"预设方案: {filter_desc.get('preset_name', '自定义')}",
        f"基础筛选: {filter_desc.get('base_filter_desc', 'N/A')}",
        f"维度筛选: {filter_desc.get('dim_filter_desc', '无')}",
        f"排序: {filter_desc.get('sort_desc', '默认')}",
        f"中英对照: {'是' if filter_desc.get('translated') else '否'}",
        f"筛选结果: {len(data_rows)} 条",
    ]

    for i, line in enumerate(info_lines):
        ws.cell(row=2 + i, column=1, value=line)

    header_row = 2 + len(info_lines) + 1  # 留一行空行

    # 表头
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for col_idx, col_name in enumerate(header, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 数据行（交替着色）
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    fill_a = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    fill_b = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    prev_pub = None
    color_toggle = False

    for row_idx, row_data in enumerate(data_rows):
        current_pub = str(row_data[COL_PUBLISHER]).strip() if row_data[COL_PUBLISHER] else "Unknown"
        if current_pub != prev_pub:
            color_toggle = not color_toggle
            prev_pub = current_pub
        row_fill = fill_a if color_toggle else fill_b

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=header_row + 1 + row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill = row_fill

    # 列宽
    col_widths = {1: 45, 2: 12, 3: 12, 4: 12, 5: 35, 6: 15, 7: 15}
    for col_idx in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_idx, 25)

    # 冻结表头
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # ── Sheet 2: 统计摘要 ──
    ws2 = wb.create_sheet("统计摘要")
    ws2.cell(row=1, column=1, value="统计摘要").font = Font(size=14, bold=True)

    r = 3

    # 基础统计
    ws2.cell(row=r, column=1, value="基础筛选统计").font = Font(size=12, bold=True)
    r += 1
    for key, label in [
        ("total_input", "输入总数"),
        ("excluded_by_source_type", "剔除(非Journal类型)"),
        ("excluded_by_language", "剔除(Chinese语言)"),
        ("after_base_filter", "基础筛选后数量"),
        ("final_count", "最终结果数"),
    ]:
        ws2.cell(row=r, column=1, value=label)
        ws2.cell(row=r, column=2, value=base_stats.get(key, "N/A"))
        r += 1

    r += 1

    # 出版商规模分布（分级统计）
    ws2.cell(row=r, column=1, value="出版商规模分布").font = Font(size=12, bold=True)
    r += 1
    tier_counts = {"≥200": 0, "100-199": 0, "50-99": 0, "20-49": 0, "10-19": 0, "5-9": 0, "2-4": 0, "1": 0}
    for _, count in pub_counter.most_common():
        if count >= 200:
            tier_counts["≥200"] += 1
        elif count >= 100:
            tier_counts["100-199"] += 1
        elif count >= 50:
            tier_counts["50-99"] += 1
        elif count >= 20:
            tier_counts["20-49"] += 1
        elif count >= 10:
            tier_counts["10-19"] += 1
        elif count >= 5:
            tier_counts["5-9"] += 1
        elif count >= 2:
            tier_counts["2-4"] += 1
        else:
            tier_counts["1"] += 1

    tier_labels = {
        "≥200": "大型（≥200本）",
        "100-199": "大型（100-199本）",
        "50-99": "中型（50-99本）",
        "20-49": "中小型（20-49本）",
        "10-19": "小型（10-19本）",
        "5-9": "微型（5-9本）",
        "2-4": "超微型（2-4本）",
        "1": "单本（仅1本）",
    }
    for tier, label in tier_labels.items():
        ws2.cell(row=r, column=1, value=label)
        ws2.cell(row=r, column=2, value=tier_counts[tier])
        r += 1
    # 出版商总数
    ws2.cell(row=r, column=1, value="合计")
    ws2.cell(row=r, column=1).font = Font(bold=True)
    ws2.cell(row=r, column=2, value=len(pub_counter))
    ws2.cell(row=r, column=2).font = Font(bold=True)
    r += 1

    r += 1

    # 出版社分布（完整列表）
    ws2.cell(row=r, column=1, value=f"出版社分布（完整列表，共 {len(pub_counter)} 家，按期刊数量降序）").font = Font(size=12, bold=True)
    r += 1
    for hdr in ["排名", "数量", "出版社"]:
        ws2.cell(row=r, column=["排名","数量","出版社"].index(hdr)+1, value=hdr).font = Font(bold=True)
    r += 1
    for rank, (pub, count) in enumerate(pub_counter.most_common(), 1):
        ws2.cell(row=r, column=1, value=rank)
        ws2.cell(row=r, column=2, value=count)
        ws2.cell(row=r, column=3, value=pub)
        r += 1

    r += 1

    # 国家/地区分布
    ws2.cell(row=r, column=1, value=f"国家/地区分布（共 {len(dist_stats['countries'])} 个）").font = Font(size=12, bold=True)
    r += 1
    for hdr in ["排名", "数量", "国家/地区"]:
        ws2.cell(row=r, column=["排名","数量","国家/地区"].index(hdr)+1, value=hdr).font = Font(bold=True)
    r += 1
    for rank, (ctry, count) in enumerate(dist_stats["countries"].most_common(), 1):
        ws2.cell(row=r, column=1, value=rank)
        ws2.cell(row=r, column=2, value=count)
        ws2.cell(row=r, column=3, value=ctry)
        r += 1

    r += 1

    # 语言分布
    ws2.cell(row=r, column=1, value=f"语言分布（共 {len(dist_stats['languages'])} 种）").font = Font(size=12, bold=True)
    r += 1
    for hdr in ["排名", "数量", "语言"]:
        ws2.cell(row=r, column=["排名","数量","语言"].index(hdr)+1, value=hdr).font = Font(bold=True)
    r += 1
    for rank, (lang, count) in enumerate(dist_stats["languages"].most_common(), 1):
        ws2.cell(row=r, column=1, value=rank)
        ws2.cell(row=r, column=2, value=count)
        ws2.cell(row=r, column=3, value=lang)
        r += 1

    r += 1

    # 学科分布
    ws2.cell(row=r, column=1, value=f"主学科分布（Subject 1，共 {len(dist_stats['subjects'])} 个）").font = Font(size=12, bold=True)
    r += 1
    for hdr in ["排名", "数量", "学科（英文）", "学科（中文）"]:
        ws2.cell(row=r, column=["排名","数量","学科（英文）","学科（中文）"].index(hdr)+1, value=hdr).font = Font(bold=True)
    r += 1
    for rank, (subj, count) in enumerate(dist_stats["subjects"].most_common(), 1):
        # 处理已翻译的 Subject（可能含 \n）
        subj_clean = subj.split("\n")[0] if "\n" in subj else subj
        zh = SUBJECT_ZH_MAP.get(subj_clean, "[未翻译]")
        ws2.cell(row=r, column=1, value=rank)
        ws2.cell(row=r, column=2, value=count)
        ws2.cell(row=r, column=3, value=subj_clean)
        ws2.cell(row=r, column=4, value=zh)
        r += 1

    # 列宽
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 65
    ws2.column_dimensions["D"].width = 40

    wb.save(filepath)
    return filepath


# ============================================================
# 预设管理
# ============================================================

def load_preset(preset_path):
    """加载 JSON 预设配置"""
    if not os.path.isabs(preset_path):
        preset_path = os.path.join(SCRIPT_DIR, preset_path)

    with open(preset_path, "r", encoding="utf-8") as f:
        config = json.load(f)

    return config


def list_presets():
    """列出 presets 目录下所有预设"""
    preset_dir = os.path.join(SCRIPT_DIR, "presets")
    if not os.path.isdir(preset_dir):
        return []

    presets = []
    for fname in sorted(os.listdir(preset_dir)):
        if fname.endswith(".json"):
            fpath = os.path.join(preset_dir, fname)
            try:
                with open(fpath, "r", encoding="utf-8") as f:
                    data = json.load(f)
                presets.append({
                    "file": fname,
                    "name": data.get("preset_name", fname),
                    "desc": data.get("preset_description", ""),
                })
            except Exception:
                pass
    return presets


def build_config_from_args(args):
    """从命令行参数构建筛选配置"""
    config = {
        "preset_name": "命令行自定义",
        "preset_description": "通过命令行参数构建",
        "base_filter": {
            "source_type": ["Journal"],
            "exclude_language_keywords": ["CHINESE"],
        },
        "dimensional_filter": {},
        "sorting": {
            "primary": {"field": "Publisher", "order": "desc", "sort_by": "count"},
            "secondary": {"field": "Source title", "order": "asc", "sort_by": "alphabetical"},
        },
        "output": {"bilingual": not args.no_translate},
    }

    if args.subject:
        config["dimensional_filter"]["subject_all"] = {
            "mode": "include", "match_type": "fuzzy", "values": [args.subject]
        }
    if args.publisher:
        config["dimensional_filter"]["publisher"] = {
            "mode": "include", "match_type": "fuzzy", "values": [args.publisher]
        }
    if args.country:
        config["dimensional_filter"]["country_region"] = {
            "mode": "include", "match_type": "exact", "values": [args.country]
        }
    if args.language:
        config["dimensional_filter"]["language"] = {
            "mode": "include", "match_type": "exact", "values": [args.language]
        }

    return config


def describe_filter(config):
    """生成筛选条件描述文本"""
    base = config.get("base_filter", {})
    dims = config.get("dimensional_filter", {})

    parts = []
    parts.append(f"Source type={','.join(base.get('source_type', []))}")
    if base.get("exclude_language_keywords"):
        parts.append(f"排除{','.join(base['exclude_language_keywords'])}")

    base_desc = "; ".join(parts)

    dim_parts = []
    for dim_name, dim_cfg in dims.items():
        if dim_cfg.get("values"):
            mode = dim_cfg.get("mode", "include")
            match = dim_cfg.get("match_type", "fuzzy")
            vals = ", ".join(dim_cfg["values"][:3])
            if len(dim_cfg["values"]) > 3:
                vals += f" 等{len(dim_cfg['values'])}项"
            dim_parts.append(f"{dim_name}({mode}/{match}): {vals}")

    dim_desc = "; ".join(dim_parts) if dim_parts else "无"

    sorting = config.get("sorting", {})
    sort_desc = f"主排序:{sorting.get('primary',{}).get('field','')} "
    sort_desc += f"{'降序' if sorting.get('primary',{}).get('order')=='desc' else '升序'}; "
    sort_desc += f"次排序:{sorting.get('secondary',{}).get('field','')}"

    return base_desc, dim_desc, sort_desc


# ============================================================
# 核心执行函数
# ============================================================

def run_filter(input_path, config, output_path=None, no_translate=False):
    """执行筛选全流程"""
    if not os.path.exists(input_path):
        print(f"[ERROR] 文件不存在: {input_path}")
        return None

    if output_path is None:
        base = os.path.splitext(input_path)[0]
        preset_name = config.get("preset_name", "filtered")
        # 清理文件名中的非法字符
        safe_name = "".join(c for c in preset_name if c not in r'\/:*?"<>|')
        output_path = os.path.join(
            os.path.dirname(base),
            f"{os.path.basename(base)}_{safe_name}.xlsx"
        )

    print(f"输入文件: {input_path}")
    print(f"输出文件: {output_path}")
    print(f"预设方案: {config.get('preset_name', 'N/A')}")
    print()

    # 1. 解析 Excel
    print("[1/6] 解析 EI Excel...")
    header, data_rows = parse_ei_excel(input_path)
    col_map = build_column_index(header)
    print(f"  表头列: {list(col_map.keys())}")
    print(f"  数据行: {len(data_rows)}")

    # 2. 筛选
    print()
    print("[2/6] 执行多维筛选...")
    filtered, base_stats = apply_filter(data_rows, col_map, config)
    print(f"  基础筛选后: {base_stats['after_base_filter']} 条")
    print(f"  最终结果: {base_stats['final_count']} 条")

    # 3. 排序
    print()
    print("[3/6] 排序...")
    sorting_config = config.get("sorting", {})
    sort_results(filtered, col_map, sorting_config)

    # 统计出版商
    pub_counter = Counter()
    for row in filtered:
        pub = str(row[COL_PUBLISHER]).strip() if row[COL_PUBLISHER] else "Unknown"
        pub_counter[pub] += 1
    print(f"  排序完成: {len(pub_counter)} 家出版社")

    # 4. 翻译
    translated = False
    print()
    print("[4/6] 中英对照翻译...")
    output_cfg = config.get("output", {})
    do_translate = output_cfg.get("bilingual", True) and not no_translate

    if do_translate:
        # Subject 翻译
        print("  [a] Subject 学科列...")
        filtered = [translate_subjects(row, col_map, SUBJECT_ZH_MAP) for row in filtered]

        # 检查未翻译
        untranslated = set()
        for row in filtered:
            for i in range(1, 9):
                idx = col_map.get(f"Subject {i}")
                if idx is not None and idx < len(row):
                    val = str(row[idx]).strip() if row[idx] else ""
                    if "[未翻译]" in val:
                        untranslated.add(val.split("\n")[0])
        if untranslated:
            print(f"  [警告] {len(untranslated)} 个学科未在映射表中")

        # Title 翻译
        print("  [b] Source title 期刊名...")
        all_titles = [str(row[COL_SOURCE_TITLE]).strip() if row[COL_SOURCE_TITLE] else "" for row in filtered]
        title_translations = translate_titles_batch(all_titles)

        for row in filtered:
            title = str(row[COL_SOURCE_TITLE]).strip() if row[COL_SOURCE_TITLE] else ""
            zh = title_translations.get(title)
            if zh:
                row[COL_SOURCE_TITLE] = f"{title}\n{zh}"

        translated = True
    else:
        print("  已跳过翻译")

    # 5. 统计
    print()
    print("[5/6] 生成统计...")
    dist_stats = generate_stats(filtered, col_map)

    # 6. 输出
    print()
    print("[6/6] 写出 Excel...")
    base_desc, dim_desc, sort_desc = describe_filter(config)
    filter_desc = {
        "source_file": os.path.basename(input_path),
        "preset_name": config.get("preset_name", "自定义"),
        "base_filter_desc": base_desc,
        "dim_filter_desc": dim_desc,
        "sort_desc": sort_desc,
        "translated": translated,
    }

    write_output(output_path, header, filtered, base_stats, dist_stats, filter_desc, pub_counter)
    print(f"  已保存: {output_path}")

    # 打印摘要
    print()
    _print_summary(base_stats, dist_stats, filter_desc, pub_counter)

    return output_path


def _print_summary(base_stats, dist_stats, filter_desc, pub_counter):
    """控制台打印统计摘要"""
    print("=" * 70)
    print("EI Compendex 期刊筛选结果摘要")
    print("=" * 70)
    print()
    print("【基础筛选】")
    print(f"  输入总数:          {base_stats['total_input']}")
    print(f"  剔除非Journal:     {base_stats['excluded_by_source_type']}")
    print(f"  剔除Chinese语言:   {base_stats['excluded_by_language']}")
    print(f"  基础筛选后:        {base_stats['after_base_filter']}")
    print(f"  最终结果:          {base_stats['final_count']}")

    print()
    print(f"【出版社分布 Top 15（共 {len(pub_counter)} 家）】")
    for rank, (pub, count) in enumerate(pub_counter.most_common(15), 1):
        print(f"  {rank:>3}. {count:>4}  {pub}")

    print()
    print(f"【国家/地区分布 Top 10（共 {len(dist_stats['countries'])} 个）】")
    for ctry, count in dist_stats["countries"].most_common(10):
        print(f"  {count:>4}  {ctry}")

    print()
    print(f"【语言分布 Top 5（共 {len(dist_stats['languages'])} 种）】")
    for lang, count in dist_stats["languages"].most_common(5):
        print(f"  {count:>4}  {lang}")

    print()
    print(f"【主学科分布 Top 15（共 {len(dist_stats['subjects'])} 个）】")
    for subj, count in dist_stats["subjects"].most_common(15):
        subj_clean = subj.split("\n")[0] if "\n" in subj else subj
        zh = SUBJECT_ZH_MAP.get(subj_clean, "")
        print(f"  {count:>4}  {subj_clean}" + (f" ({zh})" if zh else ""))

    print()
    print("=" * 70)


# ============================================================
# 统计命令
# ============================================================

def cmd_stats(input_path):
    """查看 EI 列表统计信息"""
    print(f"正在分析: {input_path}")
    print()

    header, data_rows = parse_ei_excel(input_path)
    col_map = build_column_index(header)

    print(f"Sheet: {TARGET_SHEET}")
    print(f"总条目: {len(data_rows)}")
    print(f"列数: {len(header)}")
    print(f"列名: {list(col_map.keys())}")
    print()

    # Source type
    st_counter = Counter()
    for row in data_rows:
        st = str(row[COL_SOURCE_TYPE]).strip() if row[COL_SOURCE_TYPE] else "Unknown"
        st_counter[st] += 1

    print("【Source Type 分布】")
    for st, count in st_counter.most_common():
        print(f"  {count:>5}  {st}")

    # Publisher Top 20
    pub_counter = Counter()
    for row in data_rows:
        pub = str(row[COL_PUBLISHER]).strip() if row[COL_PUBLISHER] else "Unknown"
        pub_counter[pub] += 1

    print()
    print(f"【出版社分布 Top 20（共 {len(pub_counter)} 家）】")
    for rank, (pub, count) in enumerate(pub_counter.most_common(20), 1):
        print(f"  {rank:>3}. {count:>4}  {pub}")

    # Country Top 20
    country_counter = Counter()
    for row in data_rows:
        c = str(row[COL_COUNTRY]).strip() if row[COL_COUNTRY] else "Unknown"
        country_counter[c] += 1

    print()
    print(f"【国家/地区分布 Top 20（共 {len(country_counter)} 个）】")
    for rank, (c, count) in enumerate(country_counter.most_common(20), 1):
        print(f"  {rank:>3}. {count:>4}  {c}")

    # Language
    lang_counter = Counter()
    for row in data_rows:
        l = str(row[COL_LANGUAGE]).strip() if row[COL_LANGUAGE] else "Unknown"
        lang_counter[l] += 1

    print()
    print(f"【语言分布（共 {len(lang_counter)} 种）】")
    for lang, count in lang_counter.most_common():
        print(f"  {count:>5}  {lang}")

    # Subject 1 Top 20
    subj_counter = Counter()
    for row in data_rows:
        s = str(row[COL_SUBJECT_1]).strip() if row[COL_SUBJECT_1] else "Unknown"
        subj_counter[s] += 1

    print()
    print(f"【主学科分布 Top 20（共 {len(subj_counter)} 个）】")
    for rank, (s, count) in enumerate(subj_counter.most_common(20), 1):
        zh = SUBJECT_ZH_MAP.get(s, "")
        print(f"  {rank:>3}. {count:>4}  {s}" + (f" ({zh})" if zh else ""))


# ============================================================
# 交互模式
# ============================================================

def interactive_mode():
    """交互式命令行界面"""
    print("=" * 60)
    print("  EI 期刊智能筛选工具 MVP")
    print("=" * 60)
    print()

    # 选择输入文件
    inbox = os.path.join(SCRIPT_DIR, ".dumate", "inbox")
    default_input = os.path.join(inbox, "CPXSourceList_072026.xlsx")

    if os.path.exists(default_input):
        print(f"检测到 EI 列表: {default_input}")
        use_default = input("使用此文件? (Y/n): ").strip().lower()
        if use_default != "n":
            input_path = default_input
        else:
            input_path = input("请输入文件路径: ").strip().strip('"')
    else:
        input_path = input("请输入 EI Excel 文件路径: ").strip().strip('"')

    if not os.path.exists(input_path):
        print(f"[ERROR] 文件不存在: {input_path}")
        return

    print()

    # 选择操作
    while True:
        print("─" * 40)
        print("  1. 查看列表统计")
        print("  2. 使用预设筛选")
        print("  3. 自定义筛选")
        print("  4. 列出可用预设")
        print("  0. 退出")
        print("─" * 40)

        choice = input("选择操作: ").strip()

        if choice == "0":
            print("再见!")
            break
        elif choice == "1":
            cmd_stats(input_path)
        elif choice == "2":
            presets = list_presets()
            if not presets:
                print("[INFO] 未找到预设文件。请先在 presets/ 目录下创建 JSON 预设。")
                continue

            print()
            for i, p in enumerate(presets, 1):
                print(f"  {i}. {p['name']}")
                print(f"     文件: {p['file']}")
                if p["desc"]:
                    print(f"     描述: {p['desc']}")
            print()

            sel = input(f"选择预设 (1-{len(presets)}): ").strip()
            try:
                idx = int(sel) - 1
                if 0 <= idx < len(presets):
                    preset_path = os.path.join(SCRIPT_DIR, "presets", presets[idx]["file"])
                    config = load_preset(preset_path)
                    output = run_filter(input_path, config)
                    if output:
                        print(f"\n[完成] 结果已保存: {output}")
                else:
                    print("无效选择")
            except ValueError:
                print("无效输入")

        elif choice == "3":
            print()
            print("自定义筛选（留空跳过该维度）")

            subjects = input("学科关键词（多个用逗号分隔）: ").strip()
            publishers = input("出版商关键词（多个用逗号分隔）: ").strip()
            countries = input("国家/地区（多个用逗号分隔，精确匹配）: ").strip()
            languages = input("语言（多个用逗号分隔，如 ENGLISH）: ").strip()
            no_trans = input("跳过翻译? (y/N): ").strip().lower() == "y"

            config = {
                "preset_name": "交互式自定义",
                "base_filter": {
                    "source_type": ["Journal"],
                    "exclude_language_keywords": ["CHINESE"],
                },
                "dimensional_filter": {},
                "sorting": {
                    "primary": {"field": "Publisher", "order": "desc", "sort_by": "count"},
                    "secondary": {"field": "Source title", "order": "asc", "sort_by": "alphabetical"},
                },
                "output": {"bilingual": not no_trans},
            }

            if subjects:
                vals = [s.strip() for s in subjects.split(",") if s.strip()]
                config["dimensional_filter"]["subject_all"] = {
                    "mode": "include", "match_type": "fuzzy", "values": vals
                }
            if publishers:
                vals = [s.strip() for s in publishers.split(",") if s.strip()]
                config["dimensional_filter"]["publisher"] = {
                    "mode": "include", "match_type": "fuzzy", "values": vals
                }
            if countries:
                vals = [s.strip() for s in countries.split(",") if s.strip()]
                config["dimensional_filter"]["country_region"] = {
                    "mode": "include", "match_type": "exact", "values": vals
                }
            if languages:
                vals = [s.strip() for s in languages.split(",") if s.strip()]
                config["dimensional_filter"]["language"] = {
                    "mode": "include", "match_type": "exact", "values": vals
                }

            output = run_filter(input_path, config, no_translate=no_trans)
            if output:
                print(f"\n[完成] 结果已保存: {output}")

        elif choice == "4":
            presets = list_presets()
            if presets:
                print()
                for p in presets:
                    print(f"  {p['file']}")
                    print(f"    名称: {p['name']}")
                    if p["desc"]:
                        print(f"    描述: {p['desc']}")
                    print()
            else:
                print("未找到预设文件")

        print()


# ============================================================
# 命令行入口
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="EI 期刊智能筛选工具 MVP",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )

    subparsers = parser.add_subparsers(dest="command")

    # filter 子命令
    p_filter = subparsers.add_parser("filter", help="筛选 EI 列表")
    p_filter.add_argument("input", help="输入 xlsx 文件路径")
    p_filter.add_argument("output", nargs="?", default=None, help="输出 xlsx 路径")
    p_filter.add_argument("--preset", default=None, help="JSON 预设文件路径")
    p_filter.add_argument("--subject", default=None, help="学科关键词（模糊匹配 Subject 1-8）")
    p_filter.add_argument("--publisher", default=None, help="出版商关键词（模糊匹配）")
    p_filter.add_argument("--country", default=None, help="国家/地区（精确匹配）")
    p_filter.add_argument("--language", default=None, help="语言（精确匹配，如 ENGLISH）")
    p_filter.add_argument("--no-translate", action="store_true", help="跳过翻译")

    # stats 子命令
    p_stats = subparsers.add_parser("stats", help="查看 EI 列表统计")
    p_stats.add_argument("input", help="输入 xlsx 文件路径")

    # presets 子命令
    subparsers.add_parser("presets", help="列出可用预设")

    args = parser.parse_args()

    if args.command == "filter":
        if args.preset:
            config = load_preset(args.preset)
        else:
            config = build_config_from_args(args)

        output = run_filter(args.input, config, args.output, args.no_translate)
        if output:
            print(f"\n[完成] 结果已保存: {output}")

    elif args.command == "stats":
        cmd_stats(args.input)

    elif args.command == "presets":
        presets = list_presets()
        if presets:
            for p in presets:
                print(f"  {p['file']}")
                print(f"    名称: {p['name']}")
                if p["desc"]:
                    print(f"    描述: {p['desc']}")
                print()
        else:
            print("未找到预设文件")

    else:
        # 无子命令 → 交互模式
        interactive_mode()


if __name__ == "__main__":
    main()
